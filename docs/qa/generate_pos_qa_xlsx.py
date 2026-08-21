#!/usr/bin/env python3
"""Genera docs/qa/POS_FE_QA_ANALYSIS.xlsx a partir de los mismos datos del
documento POS_FE_QA_ANALYSIS.md. Mantener AMBOS sincronizados: si cambia el
análisis en el .md, actualizar las tablas aquí y re-ejecutar:

    python3 docs/qa/generate_pos_qa_xlsx.py    # requiere openpyxl

Una hoja por módulo + Resumen + Hallazgos (regla en el CLAUDE.md raíz).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "POS_FE_QA_ANALYSIS.xlsx"

META = {
    "fecha": "2026-08-06 (actualización de confiabilidad POS)",
    "commit": "working tree (drawer/offline/idempotency audit)",
    "repo": "chepelcr/tsuru-pos-system",
    "url": "https://app.tsuru.jcampos.dev",
}

COLS = ["Funcionalidad", "Descripción", "Estado", "Incidencia (TSR)", "Bugs / Riesgos", "Prioridad QA"]

ESTADOS = {
    "✅ Completo": "Implementado en FE y conectado a un endpoint verificado en producción",
    "🟡 Parcial": "Funciona el flujo principal, pero hay partes pendientes o dependencias BE incompletas",
    "⚠️ Sin verificar": "FE implementado pero el contrato del endpoint NO está confirmado (TODO(verify-endpoint))",
    "🔒 Bloqueado por backend": "FE listo; el backend correspondiente es stub o no existe",
    "❌ Faltante": "No implementado, no accesible, o código muerto",
}

STATUS_FILL = {
    "✅": "C6EFCE",  # verde
    "🟡": "FFEB9C",  # amarillo
    "⚠️": "FFD8B0",  # naranja
    "🔒": "D9D2E9",  # morado
    "❌": "FFC7CE",  # rojo
}

# (hoja, [filas]) — cada fila sigue COLS
MODULES = [
    ("Autenticación", [
        ["Login / Logout", "Cognito vía Amplify (AuthContext); guarda redirectAfterLogin", "✅ Completo", "TSR-008", "Credenciales ya no se guardan en sessionStorage; elimina cualquier valor legacy", "Alta"],
        ["Registro", "Multi-paso con ubicación estructurada (cascada CR) + panel Sibö/cacao (TSR-113)", "✅ Completo", "TSR-008", "Verificación usa autoSignIn de Amplify sin persistir password", "Alta"],
        ["Verificación de email", "OTP Cognito; requerida antes de entrar", "✅ Completo", "—", "SES en sandbox: correos solo a direcciones verificadas (roadmap §7 paso 2)", "Alta"],
        ["Recuperar/restablecer contraseña", "Flujo Cognito completo", "✅ Completo", "—", "—", "Media"],
        ["Invitaciones (/join/:token)", "Aceptar invitación a organización", "🟡 Parcial", "TSR-012", "Botón 'Rechazar' es no-op; redirects ?redirect= se pierden", "Media"],
        ["Selección/creación de org", "SelectOrganization, CreateOrganization (onboarding)", "✅ Completo", "TSR-117", "Re-link de org por email tras pool nuevo de Cognito (verificar)", "Media"],
    ]),
    ("Dashboard", [
        ["Widgets de inicio", "SalesChart, LiveStationsPanel, TopProductsPanel con datos reales de GET …/dashboard", "✅ Completo", "—", "—", "Media"],
        ["Acciones rápidas de documento", "Crear factura/tiquete, ver documentos (RBAC doc-type)", "✅ Completo", "TSR-109", "—", "Media"],
        ["Compartir QR", "QrShareModal con URL del storefront", "✅ Completo", "—", "—", "Baja"],
    ]),
    ("POS-Documentos", [
        ["Carrito + catálogo", "ProductGrid/ProductsPanel, CartSidebar, Zustand persistido, sync offline (Dexie)", "✅ Completo", "TSR-122", "Carrito compartido entre tabs de documento (TASKS.md T7.4-futuro)", "Alta"],
        ["Detalle de línea", "LineDetailDrawer: impuestos Hacienda v4.4 (01–12/99, CABYS) y descuentos", "✅ Completo", "TSR-004", "Motor TS espejo del Python; deriva = facturas legalmente incorrectas", "Alta"],
        ["6 tipos de documento", "FE(01), TE(04), NC(03), ND(02), FC(08), FExp(09) — submódulo RBAC de creación por tipo (documents/fe…fexp)", "✅ Completo", "TSR-109", "Cajero restringido debe ver SOLO FE+TE en los 3 menús de creación", "Alta"],
        ["Checkout — sección Documento", "Condición de venta, actividad económica registrada de la org, plazo de crédito, moneda", "✅ Completo", "TSR-006", "Auto-selecciona primera actividad activa; bloquea envío sin actividad configurada", "Alta"],
        ["Checkout — sección Pagos", "Pagos múltiples: efectivo 01, tarjeta 02, cheque 03, transferencia 04, SINPE 06, otro 99", "✅ Completo", "—", "Σ pagos == total NO se valida en BE — probar pagos que no cuadran", "Alta"],
        ["Checkout — sección Receptor", "ReceiverPicker + borrador con cascada de ubicación CR (neighborhood_id → nombre)", "✅ Completo", "—", "Receptor obligatorio según tipo de doc (FE sí, TE no) — probar matriz", "Alta"],
        ["Checkout — Referencias y Copias", "ReferencesSection (obligatoria para NC/ND) + CopiesSection (correos en copia)", "🟡 Parcial", "TSR-126", "Referencias sin auditoría Nota 10/10.1", "Alta"],
        ["Multi-moneda del documento", "Precios base CRC; doc en USD/EUR divide líneas y totales por exchange_rate (UI + payload)", "✅ Completo", "—", "Redondeo tras la división — comparar totales FE vs BE en moneda extranjera", "Alta"],
        ["Venta offline", "Persiste primero en IndexedDB, intenta POST inmediato y reenvía desde la app autenticada al reconectar/iniciar sesión", "✅ Completo", "TSR-130", "Queue con estados/reintentos + token fresco; Idempotency-Key compartida con sales-be. Probar reconexión real y errores permanentes", "Alta"],
        ["Recibo de checkout", "Resultado confirmado o encolado; conserva resumen y tab hasta iniciar Nueva venta", "✅ Completo", "TSR-131", "Snapshot de total/items previo al clear; probar cierre por botón, Escape y receipt queued", "Alta"],
        ["Cierre de caja", "ClosingFlow.tsx (450 líneas, BE con soporte completo)", "❌ Faltante", "TSR-005", "Componente nunca montado — el ciclo de cajero no se puede cerrar desde la UI", "Alta"],
        ["Apertura de sesión / turno", "Selección de sesión activa; turno server-side pendiente", "🟡 Parcial", "TSR-005", "Turno server-side al iniciar sesión no implementado", "Alta"],
    ]),
    ("Facturación-Hacienda", [
        ["Lista/detalle de documentos", "Emitidos y recibidos, badges ATV, 'Pendiente' para PDF/XML nulos", "✅ Completo", "—", "—", "Alta"],
        ["Emisión (pipeline completo)", "clave/consecutivo → XML → XAdES → ATV → poll → PDF → correo", "🟡 Parcial", "TSR-001, TSR-007", "x-user-id spoofable en sales-api (default 'anonymous') — auditoría de docs legales", "Alta"],
        ["Regenerar XML / reenviar notificación", "DocumentActionModal: PDF, descarga, validación, reenviar, aceptar/rechazar", "🔒 Bloqueado por backend", "TSR-009", "Los dos endpoints jbiller son stubs — el botón existe, el BE no hace nada real", "Alta"],
        ["Aceptación de docs recibidos", "Aceptar / parcial / rechazar con envío a Hacienda", "✅ Completo", "TSR-119", "Acciones y estados traducidos ES/EN", "Alta"],
        ["Búsqueda compleja / filtros", "ComplexSearchModal, DocumentTypesFilter, fechas/tipos/montos", "🟡 Parcial", "—", "Contrato de filtros del sales-api WIP; FE envía filtros que el parser BE no soporta", "Alta"],
        ["Credenciales fiscales de la org", "P12/PIN/ATV (tarjeta Hacienda en org-settings)", "✅ Completo", "TSR-002, TSR-003", "CRÍTICO legal: credenciales en claro en BD y en GET; realm OAuth hardcodeado a staging", "Alta"],
    ]),
    ("Productos", [
        ["CRUD de productos", "ProductDrawerForm con 11 secciones fiscales (CABYS, IVA, descuentos, empaques, inventario)", "✅ Completo", "—", "Verificar mismo TaxCalculationService que LineDetailDrawer (TASKS.md T7.3)", "Alta"],
        ["Importación Excel/CSV", "ProductExcelUpload carga masiva", "⚠️ Sin verificar", "—", "TODO(verify-endpoint) en ProductExcelUpload.tsx:95", "Media"],
        ["Acciones masivas", "ProductBulkBar: activar/desactivar + seleccionar todo", "✅ Completo", "—", "—", "Media"],
    ]),
    ("Categorías", [
        ["CRUD de categorías", "Página standalone con drawer (antes solo dropdown de lectura)", "✅ Completo", "—", "Escrituras POST/PUT/DELETE con TODO(verify-endpoint) en useCategories.ts", "Media"],
    ]),
    ("Clientes-B2B", [
        ["CRUD de clientes", "Lista + detalle (Resumen/Pedidos/Tiendas/Departamentos), notas, WhatsApp", "✅ Completo", "—", "—", "Media"],
        ["Tiendas y departamentos", "ClientStoresList, ClientDepartmentsList (capa B2B)", "⚠️ Sin verificar", "—", "useStores.ts y useDepartments.ts con TODO(verify-endpoint)", "Media"],
        ["Historial de pedidos", "ClientOrderHistory real", "⚠️ Sin verificar", "—", "useClientOrders.ts sigue sin verificar; llaves comingSoon* eliminadas", "Media"],
    ]),
    ("Pedidos", [
        ["Lista con búsqueda y paginación", "OrdersPage: búsqueda + paginación normalizada (tolera snake_case y camelCase del BE)", "✅ Completo", "—", "—", "Media"],
        ["Filtros avanzados", "OrdersFiltersModal: multi-estado (default excluye entregados+cancelados), doble rango de fechas, sort 6 vías", "🟡 Parcial", "TSR-013", "String compuesto 'search' — TODO(verify-endpoint): sin confirmar que el BE lo parsee (useOrders.ts:121)", "Alta"],
        ["Detalle del pedido", "OrderDetailPage: timeline de estados, líneas, cliente/proveedor/envío, totales", "✅ Completo", "—", "Fechas DD/MM/YYYY string — probar malformadas", "Media"],
        ["Transiciones de estado", "pending→processing→shipped→delivered/cancelled vía PATCH con código numérico 1–5", "⚠️ Sin verificar", "TSR-013", "Legacy usaba PUT /status, POS usa PATCH — contrato sin confirmar (useOrders.ts:174)", "Alta"],
        ["Importación Excel de pedidos", "OrderExcelUpload + XlsxDropZone → base64 POST /orders/parse", "⚠️ Sin verificar", "TSR-013", "TODO(verify-endpoint) en shape {data,name,contentType} (useOrders.ts:209)", "Alta"],
        ["Reprocesar pedido", "ReprocessDialog + ReportColorSelector → POST /orders/{doc}/reprocess {color}", "⚠️ Sin verificar", "TSR-013", "TODO(verify-endpoint) (useOrders.ts:194)", "Alta"],
        ["Cross-docking (pedidos tipo 73)", "CrossdockingUploadDialog (Excel+color) → crossdocking/parse; Summaries (ítems/cajas/puntos); PDFPreview (iframe)", "⚠️ Sin verificar", "TSR-013", "TODO(verify-endpoint) (useOrders.ts:226); tipo 73 = código mágico del dominio", "Alta"],
        ["Adjuntos del pedido", "Descargas: PDF del pedido, Excel original, 'nuevo reporte'", "✅ Completo", "—", "Probar adjuntos nulos (pedido sin reprocesar)", "Media"],
    ]),
    ("Confirmaciones", [
        ["Lista / detalle", "ConfirmationsPage + ConfirmationDetailPage (ConfirmationCard)", "⚠️ Sin verificar", "TSR-013", "Todo useConfirmations.ts con TODO(verify-endpoint)", "Alta"],
        ["Crear confirmación", "CreateConfirmationDialog + OrderMultiPicker (selección múltiple de pedidos)", "⚠️ Sin verificar", "TSR-013", "—", "Alta"],
        ["Agregar pedidos a confirmación", "AddOrdersDialog → PUT", "⚠️ Sin verificar", "TSR-013", "—", "Alta"],
        ["Cambiar estado de confirmación", "Reusa la máquina de estados de pedidos", "⚠️ Sin verificar", "TSR-013", "—", "Alta"],
        ["Quitar pedido de confirmación", "DELETE (tolera respuesta 204 vacía)", "⚠️ Sin verificar", "TSR-013", "—", "Alta"],
        ["Correos de entrega", "Notificación por cliente al confirmar", "🔒 Bloqueado por backend", "TSR-013", "BE con branding 'Modas Laura' y EMAIL_RECIPIENT estático (email_service.py)", "Alta"],
    ]),
    ("Reportes", [
        ["Reportes de sesión/ventas", "Reporte básico por sesión (ReportePage)", "✅ Completo", "—", "—", "Media"],
        ["Analytics", "Consumidores de orgPath('/analytics') existen", "⚠️ Sin verificar", "TSR-014", "Endpoints de analytics en markets-api sin verificar", "Media"],
    ]),
    ("Sesiones-Puestos", [
        ["Sesiones", "SessionsPage + SessionConfig (creación multi-paso)", "✅ Completo", "—", "Tormenta de refetch ya corregida (cache 5 min) — verificar que no regresó", "Media"],
        ["Puestos (sucursales/terminales)", "PuestosPage, catálogo de tipos de sucursal", "🟡 Parcial", "—", "useBranchTypes con TODO — catálogo en proceso en cross-app-be", "Media"],
        ["Aprobación de cierres", "Autorización de cierre de caja", "🔒 Bloqueado por backend", "TSR-010", "is_manager default True en BE — cualquier usuario aprueba cierres", "Alta"],
    ]),
    ("Roles-RBAC", [
        ["Gestión de miembros", "Invitar, asignar rol, remover", "✅ Completo", "TSR-024", "—", "Media"],
        ["Roles y matriz de permisos", "RolesPage + RoleDrawerForm + PermissionMatrix org-scoped; catálogo espejo 1:1 del sidebar", "✅ Completo", "TSR-037", "—", "Media"],
        ["Gating de acciones app-wide", "133 elementos gateados (hide-not-disable) + nav + doc-types + org-settings por tarjeta", "✅ Completo", "TSR-110, TSR-038, TSR-109, TSR-107", "usePermissions() fail-open mientras RBAC_ENFORCEMENT=log — no confiable hasta el flip (TSR-027)", "Alta"],
    ]),
    ("Config-Organización", [
        ["9 tarjetas de configuración", "Hub + sub-páginas con RBAC por tarjeta; falso 'Pendiente' resuelto", "✅ Completo", "TSR-107, TSR-108", "Endpoints puntuales de storefront-settings con TODO(verify-endpoint) (migración 05)", "Media"],
        ["Configuración Hacienda", "Stepper de credenciales fiscales", "✅ Completo", "TSR-002", "Credenciales en claro (crítico legal) — ver hoja Facturación", "Alta"],
        ["Tema por organización", "Persistencia en organization_settings.theme", "✅ Completo", "TSR-108", "—", "Baja"],
    ]),
    ("CMS-Storefront", [
        ["Editor de contenido (CMS)", "Páginas/secciones del storefront (useCmsContent)", "⚠️ Sin verificar", "TSR-129", "RIESGO MÁS ALTO: migración 04 lo marcó 'módulo más riesgoso'; endpoints nunca ejercitados por el POS", "Alta"],
        ["Galería de medios", "Registro de media server-backed (useMediaLibrary)", "🟡 Parcial", "—", "Relativamente nuevo; probar subida/borrado", "Media"],
        ["Galería de plantillas", "Aplicar plantilla vía PUT …/organizations/{org}/template idempotente ({templateId, includeCategories}; null = desasignar)", "✅ Completo", "—", "Endpoint dedicado verificado (2026-06-13)", "Media"],
        ["Pipeline de publicación del POS", "useDeployments (markets-api): pre-deployment ready → publish → historial poll 5s (building/uploading/success/error)", "⚠️ Sin verificar", "TSR-129", "'None of these endpoints are currently exercised by POS'; compite con el pipeline real de sitios org", "Alta"],
        ["Despliegue real de sitios org (provisioner)", "Backend sales-be (TSR-118 W2, vivo 2026-07-03): SNS ORGANIZATION_REGISTERED/TEMPLATE_UPDATED → espejo del bundle → config.json live → {subdomain}.stores.tsuru.jcampos.dev", "🟡 Parcial", "TSR-118, TSR-129", "Sin UI en el POS (ni trigger ni estado); bundles de ejemplo predatan W1/W4 — sitio org renderiza shell pero no carga datos hasta rebuild con firma guest. QA E2E: registrar org → sitio vivo → datos", "Alta"],
    ]),
    ("Programas", [
        ["Programas (type=program)", "Submódulo template-gated (TSR-118 W12)", "🟡 Parcial", "TSR-118", "Filtro type puede no existir en BE; la página degrada a estado vacío", "Media"],
    ]),
    ("Perfil", [
        ["Editar perfil", "Nombre/usuario (PUT /api/users/{userId}/profile); contraseña vía Cognito", "✅ Completo", "—", "useProfile.ts con TODO(verify-endpoint) puntual", "Baja"],
    ]),
    ("Shell-Transversales", [
        ["Sidebar + navegación", "DashboardSidebar con RBAC (NAV_PERMISSION), drawer móvil, toggle colapsable, rol org-scoped en footer", "✅ Completo", "TSR-038, TSR-111", "—", "Media"],
        ["Campana de notificaciones", "NotificationsBell + NotificationsContext: niveles, CTA, targeting por app, eventos silenciosos", "🟡 Parcial", "TSR-128", "100% client-side (useState), se pierde al refrescar; source:'be' existe pero ningún backend publica", "Media"],
        ["Modo oscuro", "useDarkMode (clase dark + variables CSS); toggle en AuthNavbar y DashboardHeader", "✅ Completo", "—", "Recorrer páginas clave en dark — regla: cero colores hardcodeados", "Media"],
        ["Cambio de idioma ES/EN", "useLanguageSwitch en auth + dashboard; ES default", "✅ Completo", "TSR-119", "Fugas conocidas de documentos corregidas; mantener recorrido EN", "Alta"],
        ["Offline / sincronización", "Dexie v2 + PendingSalesSyncBridge autenticado; SyncPill online/offline/syncing/pending/error", "✅ Completo", "TSR-130", "SW ya no reenvía ventas; validar reconexión, reinicio y aislamiento por usuario", "Alta"],
        ["Exportar CSV", "Botón CSV en ReportePage (generación en navegador)", "✅ Completo", "—", "Verificar escape de comillas/comas y acentos", "Media"],
        ["Exportar PDF", "'Descargar PDF' en ReportePage (html2canvas + jsPDF, A4)", "✅ Completo", "—", "Render de canvas en reportes largos", "Media"],
        ["Impresión de recibo", "PRINT_RECEIPT.md documenta botón en src/pages/pos/POSPage.tsx — archivo que ya no existe", "❌ Faltante", "TSR-127", "Único window.print está en ReportePage; Receipt.tsx del checkout no imprime — no se puede imprimir recibo de venta", "Alta"],
        ["Paginación", "Componente Pagination compartido en todas las listas", "✅ Completo", "—", "—", "Baja"],
        ["Estados de carga/error", "Skeletons por página, ErrorBox, EmptyState, PageTransition", "✅ Completo", "—", "—", "Baja"],
        ["Responsive / móvil", "Drawers móviles, panel de notificaciones centrado, toolbar container-queries", "✅ Completo", "TSR-111, TSR-132", "Drawers/modales usan portal + overlay stack; probar página scrolleada y overlays anidados", "Media"],
        ["Sesión y token", "Amplify/Cognito refresh automático; getToken() en los 4 clientes API", "✅ Completo", "—", "Probar expiración de sesión larga (>1h abierta)", "Media"],
    ]),
    ("Motor-Fiscal", [
        ["Los 10 tipos de impuesto", "01 IVA, 02 ISC, 03 IUC, 04 ISEBA, 05 ISEBEC, 06 IPT, 07 IVACE, 08 IVARBU, 12 ISEC (5% fijo), 99 Otros — cada uno con su fórmula (taxCalculationService.ts)", "✅ Completo", "TSR-004", "Motor TS espejo del Python sin contract tests — deriva = facturas incorrectas", "Alta"],
        ["IVACE (07) base manual", "Base imponible manual con validador base_amount ≥ subtotal tras descuento (IvaTaxSection)", "✅ Completo", "—", "—", "Alta"],
        ["IVARBU (08) por factor", "factor × subtotal (catálogo de factores)", "✅ Completo", "—", "—", "Alta"],
        ["Impuestos de monto específico (03/04/05/06)", "Requieren tax_amount_id + quantity (+ percentage/volume_consumption); montos de useAllTaxAmounts aplanados a TaxAmountsById", "✅ Completo", "—", "Dos rutas de resolución (monto inline al seleccionar vs catálogo) — probar ambas", "Alta"],
        ["ISEBEC (05) por prefijo CABYS", "3401* (alcohólicas: tarifa por % alcohol) vs 2202* (no alcohólicas: monto manual con volume_consumption)", "✅ Completo", "—", "Branching por prefijo CABYS — caso de borde clave", "Alta"],
        ["Impuesto asumido por fábrica", "Routing ISEC/IUC vía has_factory_tax; otros especiales vía config estática; IVA absorbido con descuentos 01/03 (Nota 20)", "✅ Completo", "—", "Lógica más compleja del motor — matriz de casos obligatoria", "Alta"],
        ["IVA sobre base original", "Descuentos 01/03 y 02 (IVA-al-cliente) calculan IVA sobre subtotal PRE-descuento; export igual", "✅ Completo", "—", "—", "Alta"],
        ["CABYS → IVA automático", "useCabysSearch retorna tax_rate.percentage auto-aplicado al seleccionar", "✅ Completo", "—", "—", "Alta"],
        ["Cascada de descuentos secuencial", "Hasta 5 descuentos en cascada al remanente (NO suma de %); monto absoluto gana; clamp a [0, remanente]", "✅ Completo", "—", "—", "Alta"],
        ["Tipos de descuento 01/02/03/99", "Regalía (01), Regalía-IVA-al-cliente (02), Bonificación (03), Otro (99 exige reason, error REASON_REQUIRED)", "✅ Completo", "—", "CLAUDE.md del repo documenta el código viejo NATURE_DISCOUNT_REQUIRED (drift doc-vs-código)", "Alta"],
        ["Exoneraciones", "Bloque Exoneracion v4.4 (tipo doc 01–11, tarifa exonerada, monto)", "❌ Faltante", "TSR-124", "No existe en el form de línea; campos de exención del producto sin UI; hooks useAllExemptions/useExemptionValidation sin usar", "Alta"],
        ["Otros cargos (documento)", "Bloque OtherCharges v4.4 (códigos 01–10/99, hasta 15/doc; 04 exige tercero)", "❌ Faltante", "TSR-125", "No implementado en FE ni en BE de ventas; useAllOtherCharges sin usar", "Alta"],
        ["Referencias (NC/ND)", "ReferencesSection en checkout con useAllReferenceCodes/References", "🟡 Parcial", "TSR-126", "CALCULATION_AUDIT: 'parcial — pendiente auditoría vs Nota 10/10.1' (códigos + Razon obligatoria)", "Alta"],
        ["Suma de medios de pago", "Regla Σ MedioPago == TotalComprobante", "🔒 Bloqueado por backend", "—", "No se valida en el BE de ventas (CALCULATION_AUDIT §4) — probar pagos que no cuadran", "Alta"],
        ["Redondeo", "Decimal(18,5), round-half-up en el 6º dígito según spec", "⚠️ Sin verificar", "TSR-004", "FE opera en float de JS; verificar contra totales del BE", "Alta"],
    ]),
    ("Catálogos-DataApi", [
        ["Catálogos fiscales en línea/producto", "useAllTaxes/TaxRates/TaxFactors/TaxAmounts/DiscountTypes/FactoryTaxCharges/CabysSearch/MeasurementUnits/ProductTypes → LineDetailDrawer, TaxesTab, Iva/OtherTaxSection, FiscalInfo*, ProductDrawerForm", "✅ Completo", "—", "Si data-api cae, el form fiscal queda sin opciones — probar estados de error/carga", "Alta"],
        ["Catálogos de checkout", "useAllSaleConditions (DocumentSection), useAllPayments (PaymentSection), useAllReferenceCodes/References (ReferencesSection), useAllCurrencies", "✅ Completo", "—", "—", "Alta"],
        ["Catálogos de clientes", "useAllIdentifications/Countries/CustomerTypes en ClientFormBody/IdentitySection; cascada CR useStates/Counties/Districts/Neighborhoods", "✅ Completo", "—", "—", "Media"],
        ["Catálogos de org registrada (Hacienda)", "useTaxpayerInfo + useAllIdentifications/Countries en HaciendaInfoStep (consulta de contribuyente real)", "✅ Completo", "—", "—", "Alta"],
        ["Tipo de cambio", "useExchangeRates → ExchangeRateContext; conversión en ProductGrid, CartRow/Sidebar/LineEditor, Receipt, CheckoutDrawer", "✅ Completo", "—", "Roto hasta 65f2ca9 por URL default; verificar ₡/USD/EUR en cada superficie", "Alta"],
        ["Versiones de documento", "useAllDocumentVersions → DocumentVersionContext (v4.4); auto-inyección de document_version_id", "✅ Completo", "—", "No pasar document_version_id manualmente (regla del repo)", "Media"],
        ["Cache agresivo de catálogos", "staleTime 24h, gcTime 7d, persistido a localStorage (pos-system-rq-cache, maxAge 7d)", "✅ Completo", "—", "Cambio de catálogo puede tardar hasta 7 días en reflejarse: el feed de invalidación (useCatalogInvalidationFeed) está montado pero el servicio BE de notificaciones NO existe aún (docs/CATALOG_INVALIDATION_TODO.md)", "Alta"],
        ["Catálogos sin consumidor", "useAllExemptions/useExemptionValidation (TSR-124), useAllOtherCharges (TSR-125), useAllDocumentTypes, PharmaceuticalForms, Regimes, NationalTaxpayerCompanies, NotificationCodes, Transactions, TaxConditions, TaxRateCodes, useDollarRate/useEuroRate", "❌ Faltante", "TSR-124, TSR-125", "Actividades ya vienen de la org registrada; los demás hooks siguen sin UI", "Media"],
    ]),
    ("Código-Huérfano", [
        ["AnalyticsPage / AssignmentsPage", "Páginas admin en src/pages/dashboard/ sin ruta ni import — inaccesibles", "❌ Faltante", "TSR-120", "Decisión de producto: migrar o borrar", "Baja"],
        ["Flujo de dispositivo src/pages/pos/*", "SessionSetup/InventoryOpening/Payment/Success — documentado pero no cableado en Routes.tsx", "❌ Faltante", "TSR-120", "—", "Baja"],
        ["ClosingFlow.tsx", "450 líneas de cierre de caja con BE listo, nunca montado", "❌ Faltante", "TSR-005", "Funcionalidad faltante, no solo código muerto", "Alta"],
    ]),
]

HALLAZGOS = [
    ["TSR-119", "Fugas de i18n: español hardcodeado en componentes de documentos",
     "Resuelto 2026-08-06: acciones, estados, filtros y ordenamiento usan llaves ES/EN; llaves muertas clients.orders.comingSoon* eliminadas", "Media"],
    ["TSR-120", "Páginas huérfanas sin decisión de producto",
     "AnalyticsPage.tsx y AssignmentsPage.tsx sin ruta; flujo cajero src/pages/pos/* sin cablear. Decidir: migrar o eliminar", "Baja"],
    ["TSR-121", "Higiene de repo: .env.example y README obsoletos",
     ".env.example apunta a markets-api.jcampos.dev / orders-api.jcampos.dev (pre-rebrand) y nombra 'JMarkets POS'; README.md es boilerplate npm en un repo pnpm", "Baja"],
    ["TSR-122", "Carrito compartido entre tabs de documento",
     "El estado del carrito no se persiste/restaura por tab (TASKS.md T7.4-futuro); editar dos documentos en paralelo mezcla líneas", "Media"],
    ["TSR-123", "Documento de análisis QA (md + xlsx)", "Entregable de preparación de QA — Done", "—"],
    ["TSR-124", "Exoneraciones v4.4 sin UI",
     "El bloque Exoneracion (tipo doc 01–11, tarifa, monto) no existe en el form de línea; campos de exención del producto con round-trip de tipos pero sin editor (CALCULATION_AUDIT gap #8); hooks useAllExemptions/useExemptionValidation/useAllExemptionIssuingInstitutions sin usar", "Alta"],
    ["TSR-125", "Otros cargos (documento) no implementado",
     "Bloque OtherCharges v4.4 (códigos 01–10/99, hasta 15/doc) ausente en FE y en el BE de ventas; useAllOtherCharges sin usar; además Σ MedioPago == TotalComprobante no se valida en BE", "Alta"],
    ["TSR-126", "Referencias NC/ND parciales",
     "ReferencesSection sin auditoría contra Nota 10/10.1 (cobertura completa de ReferenceCode + Razon obligatoria) — riesgo de notas de crédito/débito rechazadas", "Alta"],
    ["TSR-127", "Impresión de recibo de venta inalcanzable",
     "PRINT_RECEIPT.md referencia src/pages/pos/POSPage.tsx (ya no existe); único window.print en ReportePage; Receipt.tsx del checkout no imprime — un POS que no imprime recibos", "Alta"],
    ["TSR-128", "Notificaciones sin backend ni persistencia",
     "NotificationsContext 100% client-side (useState) — se pierde al refrescar; source:'be' existe pero nada lo publica; bloquea el feed de invalidación de catálogos", "Media"],
    ["TSR-129", "Dos pipelines de despliegue sin reconciliar",
     "El POS surfacea el pipeline markets-api (pre-deployments→publish→history), nunca ejercitado; el despliegue REAL de sitios org es el provisioner de sales-be (TSR-118 W2, vivo), sin UI en el POS. Decidir cuál superficie el POS y verificar/retirar el otro", "Alta"],
    ["TSR-130", "Replay de ventas offline roto (BUG confirmado)",
     "Resuelto 2026-08-06: replay autenticado en primer plano con token fresco, queue Dexie v2, estados/reintentos, aislamiento por usuario e idempotencia org-scoped en sales-be", "Alta"],
    ["TSR-131", "Recibo de checkout perdía totales y cerraba el tab antes del resultado",
     "Resuelto 2026-08-06: snapshot previo al clear, recibo confirmado/encolado persistente y cierre del tab únicamente al iniciar Nueva venta", "Alta"],
    ["TSR-132", "Drawers/modales ocultos o recortados por scroll/transform",
     "Resuelto 2026-08-06: portal al body, overlay stack, body lock ref-counted, Escape/backdrop topmost, focus trap/restoration, dialog y 100dvh; 3 tests", "Alta"],
    ["TSR-133", "Bundle principal del POS supera 500 KB",
     "Build 2026-08-06: ~1.49 MB minificado / ~365 KB gzip. Separar rutas/chunks en un trabajo de performance independiente", "Media"],
]

RIESGOS = [
    ["R1", "Credenciales fiscales (P12, PIN, ATV) en texto plano en Postgres y retornadas por GET; realm OAuth hardcodeado a staging", "TSR-002, TSR-003", "Crítica (legal)"],
    ["R2", "Verdad fiscal duplicada: dos asignadores de consecutivo, tres motores de impuestos sincronizados 'por convención'", "TSR-004", "Crítica (legal)"],
    ["R3", "Gateway de markets-api no valida userId↔sub; el RBAC confía en una validación inexistente", "TSR-029", "Crítica (seguridad)"],
    ["R4", "Servicios Python (sales/orders/data-api) sin checks de membresía; docs fiscales y credenciales ATV expuestos", "TSR-030, TSR-031", "Crítica (seguridad)"],
]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP


def color_status(ws, col_idx):
    for row in ws.iter_rows(min_row=2):
        val = str(row[col_idx].value or "")
        for sym, color in STATUS_FILL.items():
            if val.startswith(sym):
                row[col_idx].fill = PatternFill("solid", fgColor=color)
                break


def main():
    wb = Workbook()

    # ── Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Análisis de Producto — Tsuru POS Frontend (fe/pos-system)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for k, v in [("Fecha del análisis", META["fecha"]), ("Commit analizado", META["commit"]),
                 ("Repo", META["repo"]), ("Despliegue", META["url"]),
                 ("Documento fuente", "docs/qa/POS_FE_QA_ANALYSIS.md (mantener ambos sincronizados)")]:
        ws.append([k, v])
        ws[ws.max_row][0].font = Font(bold=True)
    ws.append([])
    ws.append(["Leyenda de estados"])
    ws[ws.max_row][0].font = Font(bold=True, size=12)
    for k, v in ESTADOS.items():
        ws.append([k, v])
        for sym, color in STATUS_FILL.items():
            if k.startswith(sym):
                ws[ws.max_row][0].fill = PatternFill("solid", fgColor=color)
    ws.append([])
    ws.append(["Hojas por módulo", "Funcionalidades", "Alta prioridad"])
    ws[ws.max_row][0].font = ws[ws.max_row][1].font = ws[ws.max_row][2].font = Font(bold=True)
    for name, rows in MODULES:
        ws.append([name, len(rows), sum(1 for r in rows if r[-1] == "Alta")])
    ws.append([])
    ws.append(["Riesgos críticos (roadmap §6) que QA debe conocer"])
    ws[ws.max_row][0].font = Font(bold=True, size=12)
    ws.append(["ID", "Descripción", "TSR", "Severidad"])
    for c in ws[ws.max_row]:
        c.fill, c.font = HDR_FILL, HDR_FONT
    for r in RIESGOS:
        ws.append(r)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.alignment.wrap_text is not True:
                cell.alignment = WRAP

    # ── Una hoja por módulo ──
    for name, rows in MODULES:
        ms = wb.create_sheet(name)
        ms.append(COLS)
        for r in rows:
            ms.append(r)
        style_sheet(ms, [30, 55, 22, 20, 60, 12])
        color_status(ms, 2)

    # ── Hallazgos ──
    hs = wb.create_sheet("Hallazgos")
    hs.append(["TSR", "Hallazgo", "Detalle / evidencia", "Severidad"])
    for r in HALLAZGOS:
        hs.append(r)
    style_sheet(hs, [12, 45, 90, 12])

    wb.save(OUT)
    print(f"OK → {OUT}")


if __name__ == "__main__":
    main()
